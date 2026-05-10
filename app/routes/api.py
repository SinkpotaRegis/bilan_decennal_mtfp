from flask import Blueprint, jsonify, request, render_template, send_file, flash, redirect, url_for 
from app.models.database import db, GroupeThematique, KPI, User, Log, RoleEnum, Validation, CommentaireAnnuel, KPISAnnuel 
from sqlalchemy import text
import os
from functools import wraps
from werkzeug.utils import secure_filename
from flask_login import login_required, current_user
from datetime import datetime
from app.utils.export_utils import get_kpi_data, get_groupe_nom
from werkzeug.security import generate_password_hash
from app.decorators import log_action, admin_required
from app.kpi_mapping import KPI_META, get_kpis_pour_collecteur, get_direction_du_kpi

def collecteur_kpi_required(kpi_code):
    """
    Décorateur de restriction d'accès aux pages KPI pour les collecteurs.
 
    - Si l'utilisateur est admin ou validateur  → accès libre
    - Si l'utilisateur est collecteur           → accès uniquement si le KPI
      appartient à l'une de ses directions (User.structure)
    - Sinon                                     → 403 avec page d'erreur claire
    """
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('auth.login'))
 
            # Admin et validateur voient tout
            if current_user.role in ('admin', 'validateur'):
                return f(*args, **kwargs)
 
            # Collecteur : vérifier que le KPI lui appartient
            kpis_autorises = get_kpis_pour_collecteur(current_user.structure or '')
            if kpi_code.upper() not in kpis_autorises:
                direction_kpi = get_direction_du_kpi(kpi_code)
                flash(
                    f"Accès refusé : le KPI {kpi_code} appartient à la direction "
                    f"{direction_kpi} et ne fait pas partie de vos directions affectées "
                    f"({current_user.structure or 'aucune'}).",
                    'danger'
                )
                return redirect(url_for('api.collecteur_dashboard'))
            return f(*args, **kwargs)
        return wrapper
    return decorator

bp = Blueprint('api', __name__)

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'xlsx', 'jpg', 'png'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Tous les codes KPI valides (45 indicateurs officiels)
VALID_KPI_CODES = (
    {f'gt1pki{j:02d}' for j in range(1, 16)} |
    {f'gt2pki{j:02d}' for j in [1,2,4,5,6,7,8,9,10]} |
    {f'gt3pki{j:02d}' for j in range(1, 22)}
)
RECAP_VALIDATION_STATUSES = {'en_attente', 'valide', 'relance', 'annule'}


def _ensure_kpis_annuels_metadata_columns():
    columns = {row[1] for row in db.session.execute(text('PRAGMA table_info(kpis_annuels)')).fetchall()}

    if 'created_at' not in columns:
        db.session.execute(text('ALTER TABLE kpis_annuels ADD COLUMN created_at DATETIME'))

    if 'created_by' not in columns:
        db.session.execute(text('ALTER TABLE kpis_annuels ADD COLUMN created_by INTEGER'))

    if 'direction_concernee' not in columns:
        db.session.execute(text('ALTER TABLE kpis_annuels ADD COLUMN direction_concernee VARCHAR(200)'))

    db.session.commit()


def _normalize_validation_status(value):
    normalized = (value or 'en_attente').strip().lower()
    return normalized if normalized in RECAP_VALIDATION_STATUSES else 'en_attente'


def _load_latest_validation(kpi_code, annee):
    return db.session.execute(
        text('''
            SELECT v.statut, v.commentaire, v.created_at, u.username
            FROM validations v
            LEFT JOIN users u ON u.id = v.validateur_id
            WHERE v.kpi_code = :kpi_code AND v.annee = :annee
            ORDER BY datetime(COALESCE(v.updated_at, v.created_at)) DESC, v.id DESC
            LIMIT 1
        '''),
        {'kpi_code': kpi_code, 'annee': annee}
    ).fetchone()


def _load_commentaires_annuels(kpi_id, annee):
    commentaires = []

    commentaires_validations = db.session.execute(
        text('''
            SELECT v.commentaire, v.created_at, u.username
            FROM validations v
            LEFT JOIN users u ON u.id = v.validateur_id
            WHERE v.kpi_code = (SELECT code FROM kpis WHERE id = :kpi_id)
              AND v.annee = :annee
              AND v.commentaire IS NOT NULL
              AND TRIM(v.commentaire) != ''
            ORDER BY datetime(COALESCE(v.updated_at, v.created_at)) DESC, v.id DESC
        '''),
        {'kpi_id': kpi_id, 'annee': annee}
    ).fetchall()

    for row in commentaires_validations:
        commentaires.append({
            'commentaire': row[0] or '',
            'date_ajout': row[1].strftime('%Y-%m-%d %H:%M:%S') if hasattr(row[1], 'strftime') and row[1] else (str(row[1]) if row[1] else ''),
            '_sort_key': row[1].timestamp() if hasattr(row[1], 'timestamp') and row[1] else 0,
            'ajoute_par': row[2] or 'Validateur'
        })

    commentaires_annuels = db.session.execute(
        text('''
            SELECT commentaire, date_ajout, ajoute_par
            FROM commentaires_annuels
            WHERE kpi_id = :kpi_id AND annee = :annee
            ORDER BY id DESC
        '''),
        {'kpi_id': kpi_id, 'annee': annee}
    ).fetchall()

    for row in commentaires_annuels:
        raw_date = row[1] or ''
        sort_key = 0
        if hasattr(raw_date, 'timestamp') and raw_date:
            sort_key = raw_date.timestamp()
        else:
            try:
                sort_key = datetime.strptime(str(raw_date), '%Y-%m-%d %H:%M:%S').timestamp() if raw_date else 0
            except Exception:
                sort_key = 0

        commentaires.append({
            'commentaire': row[0] or '',
            'date_ajout': raw_date or '',
            '_sort_key': sort_key,
            'ajoute_par': row[2] or 'Utilisateur inconnu'
        })

    commentaires.sort(key=lambda item: item.get('_sort_key', 0), reverse=True)

    for commentaire in commentaires:
        commentaire.pop('_sort_key', None)

    return commentaires


def _build_recap_item(row):
    kpi_code = row.code or ''
    annee = row.annee
    validation = _load_latest_validation(kpi_code, annee)
    validation_status = _normalize_validation_status(validation[0] if validation else None)
    commentaires = _load_commentaires_annuels(row.kpi_id, annee)
    note = commentaires[0]['commentaire'] if commentaires else (validation[1] if validation and validation[1] else (row.commentaire or ''))

    return {
        'recap_id': row.recap_id,
        'kpi_id': row.kpi_id,
        'annee': annee,
        'code_indicateur': kpi_code,
        'intitule_indicateur': row.indicateur or '',
        'valeur_indicateur': row.valeur_calculee,
        'nom_collecteur': row.submitter_username or 'Utilisateur inconnu',
        'direction_concernee': row.direction_concernee or row.groupe_nom or '',
        'validation_status': validation_status,
        'note': note,
        'consult_url': url_for('api.kpi_page', kpi_code=(kpi_code or '').lower()) if kpi_code else None,
        'commentaires': commentaires,
    }


def _load_recap_rows(groupe_id, filters):
    _ensure_kpis_annuels_metadata_columns()

    # ── Construction sécurisée de la requête ─────────────────────────────────
    # On construit la clause WHERE dynamiquement avec WHERE 1=1 comme base
    params = {}
    extra_where = []

    # Filtre groupe thématique (0 ou absent = tous)
    try:
        gid = int(groupe_id) if groupe_id is not None else 0
    except (ValueError, TypeError):
        gid = 0
    if gid > 0:
        extra_where.append('k.groupe_id = :groupe_id')
        params['groupe_id'] = gid

    # Filtres texte
    if filters.get('code'):
        extra_where.append('LOWER(k.code) LIKE :code')
        params['code'] = f"%{filters['code'].strip().lower()}%"
    if filters.get('indicateur'):
        extra_where.append('LOWER(k.indicateur) LIKE :indicateur')
        params['indicateur'] = f"%{filters['indicateur'].strip().lower()}%"
    if filters.get('annee'):
        extra_where.append('ka.annee = :annee')
        params['annee'] = int(filters['annee'])

    # ── Filtre selon le rôle ──────────────────────────────────────────────────
    scope_role = current_user.role

    if scope_role == 'collecteur':
        # Le collecteur voit UNIQUEMENT les soumissions liées à sa direction
        # Logique : direction du KPI (axe_strategique) ∈ directions du collecteur
        user_dirs = [d.strip().upper() for d in (current_user.structure or '').split(',') if d.strip()]

        if user_dirs:
            # Construire la liste pour IN avec des paramètres nommés sécurisés
            dir_params = {}
            dir_placeholders = []
            for i, d in enumerate(user_dirs):
                key = f'udir{i}'
                dir_params[key] = d
                dir_placeholders.append(f':udir{i}')
            params.update(dir_params)
            params['uid'] = current_user.id

            in_clause = ', '.join(dir_placeholders)
            extra_where.append(
                f"(ka.created_by = :uid "
                f"OR UPPER(COALESCE(NULLIF(TRIM(ka.direction_concernee), ''), "
                f"NULLIF(TRIM(k.axe_strategique), ''), '')) IN ({in_clause}))"
            )
        else:
            # Aucune direction assignée → uniquement ses propres soumissions
            extra_where.append('ka.created_by = :uid')
            params['uid'] = current_user.id

    elif scope_role == 'validateur':
        # Le validateur voit les soumissions de sa propre direction
        user_dirs = [d.strip().upper() for d in (current_user.structure or '').split(',') if d.strip()]
        if user_dirs:
            dir_params = {}
            dir_placeholders = []
            for i, d in enumerate(user_dirs):
                key = f'vdir{i}'
                dir_params[key] = d
                dir_placeholders.append(f':vdir{i}')
            params.update(dir_params)
            in_clause = ', '.join(dir_placeholders)
            extra_where.append(
                f"UPPER(COALESCE(NULLIF(TRIM(ka.direction_concernee), ''), "
                f"NULLIF(TRIM(k.axe_strategique), ''), '')) IN ({in_clause})"
            )
        else:
            extra_where.append('1 = 0')
    # Admin : pas de filtre direction → voit tout

    # ── Construction de la clause WHERE finale ────────────────────────────────
    where_sql = 'WHERE 1=1'
    if extra_where:
        where_sql += ' AND ' + ' AND '.join(extra_where)

    sql = f'''
        SELECT
            ka.id              AS recap_id,
            ka.kpi_id,
            ka.annee,
            k.code,
            k.indicateur,
            ka.valeur_calculee,
            ka.commentaire,
            COALESCE(NULLIF(TRIM(ka.direction_concernee), ''),
                     NULLIF(TRIM(k.axe_strategique), ''),
                     g.nom)   AS direction_concernee,
            g.nom              AS groupe_nom,
            ka.created_at,
            ka.created_by,
            u.username         AS submitter_username
        FROM kpis_annuels ka
        JOIN kpis k             ON k.id = ka.kpi_id
        JOIN groupes_thematiques g ON g.id = k.groupe_id
        LEFT JOIN users u       ON u.id = ka.created_by
        {where_sql}
        ORDER BY
            COALESCE(datetime(ka.created_at), '1970-01-01 00:00:00') DESC,
            k.code ASC,
            ka.id DESC
    '''

    try:
        rows = db.session.execute(text(sql), params).fetchall()
    except Exception as e:
        print(f"[_load_recap_rows] SQL error: {e}")
        print(f"SQL: {sql}")
        print(f"Params: {params}")
        return []

    items = [_build_recap_item(row) for row in rows]

    # Filtre statut côté Python (après validation)
    if filters.get('statut'):
        statut = _normalize_validation_status(filters.get('statut'))
        items = [item for item in items if item['validation_status'] == statut]

    return items


def _paginate_items(items, page, per_page):
    total = len(items)
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, pages))
    start = (page - 1) * per_page
    end = start + per_page
    return {
        'items': items[start:end],
        'total': total,
        'pages': pages,
        'page': page,
    }

# ============ PAGES HTML ============
@bp.route('/rapport_codir')
@login_required
@log_action('VIEW_PAGE')
def rapport_codir():
    return render_template('rapport_codir.html', active_page='rapport_codir')


@bp.route('/')
@login_required
@log_action('VIEW_PAGE')
def dashboard():
    if current_user.role == 'collecteur':
        return redirect(url_for('api.collecteur_dashboard'))
    return render_template('dashboard.html', active_page='dashboard')


@bp.route('/collecte_avancee')
@login_required
@log_action('VIEW_PAGE')
def collecte_avancee():
    return render_template('collecte_avancee.html', active_page='collecte_avancee')


@bp.route('/graphique')
@login_required
@log_action('VIEW_PAGE')
def graphique():
    return render_template('graphique.html', active_page='graphique')

# ============ PAGES KPI ============

@bp.route('/kpi/<kpi_code>')
@login_required
@log_action('VIEW_KPI')
def kpi_page(kpi_code):
    kpi_code_lower = kpi_code.lower().strip()
    kpi_code_upper = kpi_code_lower.upper()

    # Restriction collecteur et validateur
    if current_user.role in ('collecteur', 'validateur'):
        user_dirs = [d.strip().upper() for d in (current_user.structure or '').split(',') if d.strip()]
        if user_dirs:
            # Direction du KPI : KPI_META d'abord, sinon BDD (axe_strategique)
            kpi_obj_temp = KPI.query.filter(db.func.lower(KPI.code) == kpi_code_lower).first()
            meta_temp    = KPI_META.get(kpi_code_upper, {})
            dir_kpi      = (
                meta_temp.get('direction')
                or (kpi_obj_temp.axe_strategique if kpi_obj_temp else '')
                or ''
            ).strip().upper()

            if dir_kpi and dir_kpi not in user_dirs:
                flash(
                    f"Accès refusé : le KPI {kpi_code_upper} appartient à la direction "
                    f"{dir_kpi} et ne fait pas partie de vos directions affectées "
                    f"({current_user.structure or 'aucune'}).",
                    'danger'
                )
                return redirect(url_for('api.collecteur_dashboard'))

    # Charger le KPI depuis la BDD
    kpi = KPI.query.filter(db.func.lower(KPI.code) == kpi_code_lower).first()
    if not kpi:
        from flask import abort
        abort(404)

    groupe = GroupeThematique.query.get(kpi.groupe_id)
    groupe_nom = groupe.nom if groupe else ''

    # Direction : KPI_META en priorité, sinon axe_strategique (BDD), sinon ''
    try:
        from app.kpi_mapping import DIRECTION_LABELS
        meta           = KPI_META.get(kpi_code_upper, {})
        direction_code = (
            meta.get('direction')
            or (kpi.axe_strategique or '').strip().upper()
            or ''
        )
        direction_label = DIRECTION_LABELS.get(direction_code, direction_code or 'Non définie')
    except Exception:
        direction_code  = (kpi.axe_strategique or '').strip().upper()
        direction_label = direction_code or 'Non définie'

    # Données annuelles + preuves
    rows = db.session.execute(text('''
        SELECT ka.annee, ka.numerateur_valeur, ka.denominateur_valeur,
               ka.valeur_calculee, ka.commentaire,
               COALESCE(
                   (SELECT v.statut FROM validations v
                    WHERE v.kpi_code = :code AND v.annee = ka.annee
                    ORDER BY v.id DESC LIMIT 1),
                   'en_attente'
               ) AS statut
        FROM kpis_annuels ka
        WHERE ka.kpi_id = :kpi_id
        ORDER BY ka.annee
    '''), {'kpi_id': kpi.id, 'code': kpi.code}).fetchall()

    preuves_rows = db.session.execute(text(
        'SELECT id, annee, titre, type_fichier, chemin_fichier FROM preuves_documentaires WHERE kpi_id=:kid ORDER BY annee, id'
    ), {'kid': kpi.id}).fetchall()

    preuves_par_annee = {}
    for p in preuves_rows:
        nom_fichier = p[4].split('/')[-1] if p[4] else p[2]
        preuves_par_annee.setdefault(p[1], []).append({
            'id':   p[0],
            'nom':  p[2],
            'type': p[3] or 'FILE',
            'url':  f'/static/preuves/{kpi.id}/{nom_fichier}',
        })

    donnees_bdd = {}
    for r in rows:
        annee = r[0]
        donnees_bdd[annee] = {
            'numerateur':   r[1],
            'denominateur': r[2],
            'valeur':       r[3],
            'commentaire':  r[4] or '',
            'statut':       r[5] or 'en_attente',
            'preuves':      preuves_par_annee.get(annee, []),
            'nb_preuves':   len(preuves_par_annee.get(annee, [])),
        }

    # Page dédiée existante ? (compatibilité ascendante)
    import os as _os
    template_dedie = f'kpi_{kpi_code_lower}.html'
    templates_dir  = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', 'templates')
    if _os.path.exists(_os.path.join(templates_dir, template_dedie)):
        return render_template(template_dedie)

    return render_template(
        'kpi_dynamique.html',
        kpi             = kpi,
        groupe_nom      = groupe_nom,
        direction_label = direction_label,
        direction_code  = direction_code,
        donnees_bdd     = donnees_bdd,
    )


# ============ API KPIs ============
@bp.route('/api/kpis')
@login_required
def get_kpis():
    kpis = KPI.query.all()
    return jsonify([k.to_dict() for k in kpis])


@bp.route('/api/groupes')
@login_required
def get_groupes():
    groupes = GroupeThematique.query.all()
    return jsonify([{'id': g.id, 'nom': g.nom} for g in groupes])


# ============ API DONNEES ANNUELES ============
@bp.route('/api/kpis/annuel/<int:id>')
@login_required
def get_kpi_annuel(id):
    try:
        annees = list(range(2016, 2027))

        result = db.session.execute(
            text('SELECT annee, numerateur_valeur, denominateur_valeur, valeur_calculee, commentaire FROM kpis_annuels WHERE kpi_id = :kpi_id ORDER BY annee'),
            {'kpi_id': id}
        )
        data = {}
        for row in result:
            data[row[0]] = {
                'numerateur': row[1] if row[1] is not None else 0,
                'denominateur': row[2] if row[2] is not None else 0,
                'valeur': row[3] if row[3] is not None else 0,
                'commentaire': row[4] if row[4] is not None else ''
            }

        fichiers_par_annee = {}
        fichiers_result = db.session.execute(
            text('SELECT id, annee, titre, chemin_fichier FROM preuves_documentaires WHERE kpi_id = :kpi_id ORDER BY annee, id'),
            {'kpi_id': id}
        )
        for row in fichiers_result:
            file_id, annee, titre, chemin = row
            if annee not in fichiers_par_annee:
                fichiers_par_annee[annee] = []
            fichiers_par_annee[annee].append({'id': file_id, 'nom': titre, 'chemin': chemin})

        annuel = []
        for annee in annees:
            annuel.append({
                'annee': annee,
                'numerateur': data.get(annee, {}).get('numerateur', 0),
                'denominateur': data.get(annee, {}).get('denominateur', 0),
                'valeur': data.get(annee, {}).get('valeur', None),  # None = non saisi
                'commentaire': data.get(annee, {}).get('commentaire', ''),
                'fichiers': fichiers_par_annee.get(annee, []),
                'nb_preuves': len(fichiers_par_annee.get(annee, [])),
            })
        return jsonify(annuel)
    except Exception as e:
        print(f"Erreur dans get_kpi_annuel: {e}")
        return jsonify({'error': str(e)}), 500


@bp.route('/api/kpis/annuel/update/<int:id>', methods=['POST'])
@login_required
@log_action('UPDATE_KPI')
def update_kpi_annuel(id):
    try:
        _ensure_kpis_annuels_metadata_columns()

        data = request.json
        annee = data.get('annee')
        numerateur = float(data.get('numerateur', 0) or 0)
        denominateur = float(data.get('denominateur', 0) or 0)
        commentaire = data.get('commentaire', '')
        direction_concernee = (data.get('direction') or current_user.structure or current_user.direction or '').strip()

        kpi = KPI.query.get_or_404(id)
        valeur_calculee = kpi.calculer_valeur(numerateur, denominateur)

        db.session.execute(
            text('''
                INSERT INTO kpis_annuels (
                    kpi_id,
                    annee,
                    numerateur_valeur,
                    denominateur_valeur,
                    valeur_calculee,
                    commentaire,
                    direction_concernee,
                    created_at,
                    created_by
                ) VALUES (
                    :kpi_id,
                    :annee,
                    :numerateur,
                    :denominateur,
                    :valeur,
                    :commentaire,
                    :direction_concernee,
                    :created_at,
                    :created_by
                )
                ON CONFLICT(kpi_id, annee) DO UPDATE SET
                    numerateur_valeur = excluded.numerateur_valeur,
                    denominateur_valeur = excluded.denominateur_valeur,
                    valeur_calculee = excluded.valeur_calculee,
                    commentaire = excluded.commentaire,
                    direction_concernee = COALESCE(NULLIF(excluded.direction_concernee, ''), kpis_annuels.direction_concernee),
                    created_at = COALESCE(kpis_annuels.created_at, excluded.created_at),
                    created_by = COALESCE(kpis_annuels.created_by, excluded.created_by)
            '''),
            {
                'kpi_id': id,
                'annee': annee,
                'numerateur': numerateur,
                'denominateur': denominateur,
                'valeur': valeur_calculee,
                'commentaire': commentaire,
                'direction_concernee': direction_concernee,
                'created_at': datetime.utcnow(),
                'created_by': current_user.id,
            }
        )
        db.session.commit()

        if annee == 2026:
            kpi.valeur_atteinte = valeur_calculee
            kpi.numerateur_valeur = numerateur
            kpi.denominateur_valeur = denominateur
            db.session.commit()

        # Récupérer l'id de la ligne kpis_annuels créée/mise à jour
        ligne = db.session.execute(
            text('SELECT id FROM kpis_annuels WHERE kpi_id=:kpi_id AND annee=:annee'),
            {'kpi_id': id, 'annee': annee}
        ).fetchone()
        kpi_annuel_id = ligne[0] if ligne else None

        return jsonify({'status': 'success', 'valeur': valeur_calculee, 'kpi_annuel_id': kpi_annuel_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============ EXPORT EXCEL ============
@bp.route('/api/kpis/export/excel/<int:id>')
@login_required
@log_action('EXPORT_EXCEL')
def export_kpi_excel(id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import LineChart, Reference
    from io import BytesIO

    data = get_kpi_data(id, db)
    if not data:
        return jsonify({'error': 'KPI non trouvé'}), 404

    kpi = data['kpi']
    donnees = data['annuel']
    valeurs = data['valeurs']
    annees = data['annees']
    documents = data['documents']

    kpi_dict = {
        'id': kpi[0],
        'code': kpi[3] or '',
        'indicateur': kpi[4] or '',
        'definition': kpi[5] or '',
        'objectif': kpi[6] or '',
        'valeur_cible': kpi[7] or 0,
        'unite': kpi[9] or '%',
        'source_verification': kpi[10] or '',
        'responsable': kpi[11] or '',
        'numerateur_label': kpi[16] or 'Numérateur',
        'denominateur_label': kpi[17] or 'Dénominateur',
        'frequence_collecte': kpi[14] or ''
    }

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="006633", end_color="006633", fill_type="solid")

    ws = wb.active
    ws.title = "Fiche_Indicateur"

    ws.merge_cells('A1:B1')
    ws['A1'] = "FICHE INDICATEUR - BILAN DECENNAL MTFP"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = header_fill

    groupe_nom = get_groupe_nom(kpi[2])

    infos = [
        ("Code", kpi_dict['code']),
        ("Indicateur", kpi_dict['indicateur']),
        ("Thème", groupe_nom),
        ("Définition", kpi_dict['definition'][:500]),
        ("Objectif", kpi_dict['objectif'][:500]),
        ("Formule", f"({kpi_dict['numerateur_label']} / {kpi_dict['denominateur_label']}) × 100"),
        ("Cible 2026", f"{kpi_dict['valeur_cible']} {kpi_dict['unite']}"),
        ("Responsable", kpi_dict['responsable']),
        ("Sources", kpi_dict['source_verification'][:500])
    ]

    row = 3
    for label, value in infos:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        row += 1

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 60

    ws2 = wb.create_sheet("Donnees_Annuelles")

    headers = ['Année', kpi_dict['numerateur_label'], kpi_dict['denominateur_label'], 'Taux (%)', 'Observations']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(donnees, 2):
        ws2.cell(row=row_idx, column=1, value=row[0]).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=2, value=row[1]).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=3, value=row[2]).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=4, value=round(row[3], 2)).alignment = Alignment(horizontal='center')
        ws2.cell(row=row_idx, column=5, value=row[4]).alignment = Alignment(horizontal='left')

    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 18
    ws2.column_dimensions['E'].width = 30

    if len(donnees) >= 2:
        ws3 = wb.create_sheet("Graphique")

        ws3['A1'] = "Année"
        ws3['B1'] = "Taux (%)"

        for i, row in enumerate(donnees, 2):
            ws3.cell(row=i, column=1, value=row[0])
            ws3.cell(row=i, column=2, value=round(row[3], 2))

        chart = LineChart()
        chart.title = f"Evolution de l'indicateur"
        chart.style = 13
        chart.y_axis.title = "Taux (%)"
        chart.x_axis.title = "Année"
        chart.height = 8
        chart.width = 15

        data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(donnees) + 1, max_col=2)
        categories = Reference(ws3, min_col=1, min_row=2, max_row=len(donnees) + 1)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(categories)
        ws3.add_chart(chart, "D2")

        debut = donnees[0][3]
        fin = donnees[-1][3]
        n = len(donnees) - 1
        tcac = ((fin / debut) ** (1 / n) - 1) * 100 if debut > 0 else 0
        evolution = ((fin - debut) / debut) * 100 if debut > 0 else 0

        ws3['A15'] = "STATISTIQUES"
        ws3['A15'].font = Font(bold=True, size=12)
        ws3['A16'] = f"Valeur {donnees[0][0]}: {debut:.2f}%"
        ws3['A17'] = f"Valeur {donnees[-1][0]}: {fin:.2f}%"
        ws3['A18'] = f"TCAC: {tcac:.2f}%"
        ws3['A19'] = f"Evolution totale: {evolution:+.2f}%"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f"KPI_{kpi_dict['code']}_{datetime.now().strftime('%Y%m%d')}.xlsx")


@bp.route('/api/kpis/analyse/<int:id>')
@login_required
def get_kpi_analyse(id):
    kpi = KPI.query.get_or_404(id)
    return jsonify({
        'facteurs_explicatifs': kpi.facteurs_explicatifs,
        'benchmarking': kpi.benchmarking,
        'recommandations': kpi.recommandations
    })


# ============ EXPORT PDF ============
@bp.route('/api/kpis/export/pdf/<int:id>')
@login_required
@log_action('EXPORT_PDF')
def export_kpi_pdf(id):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from io import BytesIO
    import matplotlib.pyplot as plt
    import matplotlib
    matplotlib.use('Agg')

    data = get_kpi_data(id, db)
    if not data:
        return jsonify({'error': 'KPI non trouvé'}), 404

    kpi = data['kpi']
    donnees = data['annuel']
    valeurs = data['valeurs']
    annees = data['annees']
    documents = data['documents']

    kpi_dict = {
        'id': kpi[0],
        'code': kpi[3] or '',
        'indicateur': kpi[4] or '',
        'definition': kpi[5] or '',
        'objectif': kpi[6] or '',
        'valeur_cible': kpi[7] or 0,
        'unite': kpi[9] or '%',
        'source_verification': kpi[10] or '',
        'responsable': kpi[11] or '',
        'numerateur_label': kpi[16] or 'Numérateur',
        'denominateur_label': kpi[17] or 'Dénominateur',
        'frequence_collecte': kpi[14] or ''
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#006633'), alignment=1)
    section_style = ParagraphStyle('SectionTitle', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#006633'), alignment=0)

    elements.append(Paragraph("FICHE INDICATEUR - BILAN DECENNAL MTFP", title_style))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f"{kpi_dict['code']} - {kpi_dict['indicateur']}", styles['Heading2']))
    elements.append(Spacer(1, 20))

    groupe_nom = get_groupe_nom(kpi[2])

    data_info = [
        ['Thème', groupe_nom],
        ['Définition', kpi_dict['definition'][:300]],
        ['Objectif', kpi_dict['objectif'][:300]],
        ['Formule', f"({kpi_dict['numerateur_label']} / {kpi_dict['denominateur_label']}) × 100"],
        ['Cible 2026', f"{kpi_dict['valeur_cible']} {kpi_dict['unite']}"],
        ['Sources', kpi_dict['source_verification'][:300]],
        ['Responsable', kpi_dict['responsable']],
        ['Fréquence', kpi_dict['frequence_collecte']]
    ]

    table_info = Table(data_info, colWidths=[3 * cm, 12 * cm])
    table_info.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#006633')),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('BACKGROUND', (1, 0), (1, -1), colors.HexColor('#e8f5e9')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    elements.append(table_info)
    elements.append(Spacer(1, 20))

    elements.append(PageBreak())
    elements.append(Paragraph("TABLEAU DES DONNEES ANNUELES", section_style))
    elements.append(Spacer(1, 10))

    data_table = [['Année', kpi_dict['numerateur_label'], kpi_dict['denominateur_label'], 'Taux (%)', 'Observations']]
    for row in donnees:
        data_table.append([row[0], row[1], row[2], f"{row[3]:.2f}%", (row[4] or '')[:50]])

    table_data = Table(data_table, colWidths=[1.5 * cm, 3 * cm, 3 * cm, 2.5 * cm, 5 * cm])
    table_data.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#006633')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(table_data)
    elements.append(Spacer(1, 20))

    elements.append(PageBreak())
    elements.append(Paragraph("COURBE D'EVOLUTION", section_style))
    elements.append(Spacer(1, 10))

    if len(donnees) >= 2:
        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(annees, valeurs, marker='o', linewidth=2, markersize=8, color='#006633')
        ax.set_xlabel('Année', fontsize=12)
        ax.set_ylabel('Taux (%)', fontsize=12)
        ax.set_title(f"Evolution de {kpi_dict['indicateur'][:50]}", fontsize=14, fontweight='bold')
        ax.grid(True, linestyle='--', alpha=0.7)
        ax.set_ylim(0, max(valeurs) + 10 if valeurs else 100)

        img_buffer = BytesIO()
        plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()

        img = Image(img_buffer, width=18 * cm, height=9 * cm)
        elements.append(img)
        elements.append(Spacer(1, 20))

        debut = donnees[0][3]
        fin = donnees[-1][3]
        n = len(donnees) - 1
        tcac = ((fin / debut) ** (1 / n) - 1) * 100 if debut > 0 else 0
        evolution = ((fin - debut) / debut) * 100 if debut > 0 else 0

        data_stats = [
            [f'Valeur {donnees[0][0]}', f"{debut:.2f}%"],
            [f'Valeur {donnees[-1][0]}', f"{fin:.2f}%"],
            ['TCAC', f"{tcac:.2f}%"],
            ['Évolution totale', f"{evolution:+.2f}%"]
        ]
        table_stats = Table(data_stats, colWidths=[4 * cm, 4 * cm])
        table_stats.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#006633')),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(table_stats)

    doc.build(elements)
    buffer.seek(0)

    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=f"KPI_{kpi_dict['code']}_{datetime.now().strftime('%Y%m%d')}.pdf")


# ============ PREUVES DOCUMENTAIRES ============
@bp.route('/api/preuves/upload', methods=['POST'])
@login_required
@log_action('UPLOAD_PREUVE')
def preuve_upload():
    MAX_SIZE = 50 * 1024 * 1024  # 50 Mo
    EXTS_OK  = {'pdf','doc','docx','xls','xlsx','ppt','pptx','jpg','jpeg','png','gif','mp4','zip','csv','txt'}

    if 'fichier' not in request.files:
        return jsonify({'error': 'Aucun fichier reçu'}), 400

    fichier       = request.files['fichier']
    kpi_id        = request.form.get('kpi_id')
    annee         = request.form.get('annee')
    direction     = request.form.get('direction', '')

    if not fichier or fichier.filename == '':
        return jsonify({'error': 'Fichier vide'}), 400

    ext = fichier.filename.rsplit('.', 1)[-1].lower() if '.' in fichier.filename else ''
    if ext not in EXTS_OK:
        return jsonify({'error': f'Type non autorisé : .{ext}'}), 400

    fichier.seek(0, 2); taille = fichier.tell(); fichier.seek(0)
    if taille > MAX_SIZE:
        return jsonify({'error': f'Fichier trop lourd ({taille/1024/1024:.1f} Mo, max 50 Mo)'}), 400

    nom_securise = secure_filename(fichier.filename)
    nom_unique   = f"{kpi_id}_{annee}_{current_user.id}_{nom_securise}"
    dossier      = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'app', 'static', 'preuves', str(kpi_id))
    os.makedirs(dossier, exist_ok=True)
    chemin = os.path.join(dossier, nom_unique)
    fichier.save(chemin)

    db.session.execute(text('''
        INSERT INTO preuves_documentaires (kpi_id, annee, titre, type_fichier, chemin_fichier, date_ajout, ajoute_par)
        VALUES (:kpi_id, :annee, :titre, :type, :chemin, :date_ajout, :ajoute_par)
    '''), {
        'kpi_id': int(kpi_id), 'annee': int(annee),
        'titre': nom_securise, 'type': ext.upper(),
        'chemin': chemin,
        'date_ajout': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'ajoute_par': current_user.username,
    })
    db.session.commit()

    preuve = db.session.execute(text(
        'SELECT id FROM preuves_documentaires WHERE chemin_fichier=:c'
    ), {'c': chemin}).fetchone()

    return jsonify({
        'status': 'success',
        'id':     preuve[0] if preuve else None,
        'nom':    nom_securise,
        'type':   ext.upper(),
        'url':    f'/static/preuves/{kpi_id}/{nom_unique}',
        'taille': f'{taille/1024/1024:.2f} Mo',
    })


@bp.route('/api/preuves/<int:preuve_id>', methods=['DELETE'])
@login_required
def preuve_delete(preuve_id):
    row = db.session.execute(text(
        'SELECT chemin_fichier, ajoute_par FROM preuves_documentaires WHERE id=:id'
    ), {'id': preuve_id}).fetchone()
    if not row:
        return jsonify({'error': 'Preuve introuvable'}), 404
    if current_user.role == 'collecteur' and row[1] != current_user.username:
        return jsonify({'error': 'Accès refusé'}), 403
    if row[0] and os.path.exists(row[0]):
        try: os.remove(row[0])
        except: pass
    db.session.execute(text('DELETE FROM preuves_documentaires WHERE id=:id'), {'id': preuve_id})
    db.session.commit()
    return jsonify({'status': 'success'})


# ============ UPLOAD FICHIERS ============
@bp.route('/api/kpis/upload/<int:id>', methods=['POST'])
@login_required
@log_action('UPLOAD_FILE')
def upload_kpi_file(id):
    try:
        kpi = KPI.query.get_or_404(id)

        if 'file' not in request.files:
            return jsonify({'error': 'Aucun fichier'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nom de fichier vide'}), 400

        annee = request.form.get('annee', '')

        if file and allowed_file(file.filename):
            kpi_code = kpi.code or f"KPI_{id}"
            kpi_nom = kpi.indicateur[:50].replace(' ', '_').replace('/', '_').replace('\\', '_') \
                .replace('é', 'e').replace('è', 'e').replace('ê', 'e') \
                .replace('à', 'a').replace('ô', 'o').replace('î', 'i')

            kpi_folder_name = f"{kpi_code}_{kpi_nom}"
            kpi_folder_path = os.path.join(UPLOAD_FOLDER, kpi_folder_name)
            os.makedirs(kpi_folder_path, exist_ok=True)

            extension = file.filename.rsplit('.', 1)[1].lower()
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{annee}_{kpi_code}_{timestamp}.{extension}"

            filepath = os.path.join(kpi_folder_path, filename)
            file.save(filepath)

            relative_path = f"{kpi_folder_name}/{filename}"

            result = db.session.execute(
                text('INSERT INTO preuves_documentaires (kpi_id, annee, titre, chemin_fichier, date_ajout, ajoute_par) VALUES (:kpi_id, :annee, :titre, :chemin, :date_ajout, :ajoute_par) RETURNING id'),
                {'kpi_id': id, 'annee': annee, 'titre': file.filename, 'chemin': relative_path, 'date_ajout': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'ajoute_par': current_user.username}
            )
            inserted_id = result.fetchone()[0]
            db.session.commit()

            return jsonify({'status': 'success', 'id': inserted_id, 'filename': filename, 'full_path': relative_path, 'annee': annee})
        else:
            return jsonify({'error': 'Format non autorisé'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/kpis/fichier/<int:kpi_id>/<int:annee>/<int:file_id>', methods=['DELETE'])
@login_required
@log_action('DELETE_FILE')
def delete_fichier(kpi_id, annee, file_id):
    try:
        result = db.session.execute(text('SELECT chemin_fichier FROM preuves_documentaires WHERE id = :id AND kpi_id = :kpi_id AND annee = :annee'), {'id': file_id, 'kpi_id': kpi_id, 'annee': annee})
        row = result.fetchone()
        if not row:
            return jsonify({'error': 'Fichier non trouvé'}), 404

        chemin_relatif = row[0]
        chemin_absolu = os.path.join(UPLOAD_FOLDER, chemin_relatif)

        if os.path.exists(chemin_absolu):
            os.remove(chemin_absolu)

        db.session.execute(text('DELETE FROM preuves_documentaires WHERE id = :id'), {'id': file_id})
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Fichier supprimé'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/kpis/annuel/delete/<int:kpi_id>/<int:annee>', methods=['DELETE'])
@login_required
@log_action('DELETE_KPI_DATA')
def delete_ligne_annuelle(kpi_id, annee):
    try:
        fichiers_result = db.session.execute(text('SELECT id, chemin_fichier FROM preuves_documentaires WHERE kpi_id = :kpi_id AND annee = :annee'), {'kpi_id': kpi_id, 'annee': annee})
        fichiers = fichiers_result.fetchall()

        for fichier in fichiers:
            file_id, chemin_relatif = fichier
            chemin_absolu = os.path.join(UPLOAD_FOLDER, chemin_relatif)
            if os.path.exists(chemin_absolu):
                os.remove(chemin_absolu)

        db.session.execute(text('DELETE FROM preuves_documentaires WHERE kpi_id = :kpi_id AND annee = :annee'), {'kpi_id': kpi_id, 'annee': annee})
        db.session.execute(text('DELETE FROM kpis_annuels WHERE kpi_id = :kpi_id AND annee = :annee'), {'kpi_id': kpi_id, 'annee': annee})
        db.session.commit()
        return jsonify({'status': 'success', 'message': f'Données de {annee} réinitialisées'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============ SERVIR LES FICHIERS ============
@bp.route('/uploads/<path:filepath>')
@login_required
def download_file(filepath):
    try:
        requested_path = os.path.join(UPLOAD_FOLDER, filepath)
        requested_path = os.path.normpath(requested_path)
        
        if not requested_path.startswith(os.path.normpath(UPLOAD_FOLDER)):
            return jsonify({'error': 'Accès refusé'}), 403
        
        if not os.path.exists(requested_path) or not os.path.isfile(requested_path):
            return jsonify({'error': 'Fichier non trouvé'}), 404
        
        return send_file(requested_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/fichier/existe/<path:filepath>')
@login_required
def check_file_exists(filepath):
    try:
        requested_path = os.path.join(UPLOAD_FOLDER, filepath)
        requested_path = os.path.normpath(requested_path)
        
        if not requested_path.startswith(os.path.normpath(UPLOAD_FOLDER)):
            return jsonify({'exists': False, 'error': 'Accès refusé'}), 403
        
        exists = os.path.exists(requested_path) and os.path.isfile(requested_path)
        return jsonify({'exists': exists})
    except Exception as e:
        return jsonify({'exists': False, 'error': str(e)}), 500


# ####################################################
# ##############GESTION DU COTE ADMINISTRATEUR########
# ####################################################

@bp.route('/admin/dashboard')
@login_required
@admin_required
@log_action('VIEW_ADMIN')
def admin_dashboard():
    return render_template('admin/dashboard_admin.html', active_page='admin_dashboard')


@bp.route('/admin/liste-kpis')
@login_required
@admin_required
@log_action('VIEW_ADMIN')
def admin_liste_kpis_page():
    groupes = GroupeThematique.query.order_by(GroupeThematique.nom).all()
    return render_template(
        'admin/liste_kpis.html',
        active_page='admin_liste_kpis',
        groupes=groupes,
    )


# ── Menu KPI dynamique filtré selon la direction du collecteur ───────────────
@bp.route('/api/collecteur/kpis-menu')
@login_required
def api_kpis_menu():
    """
    Retourne en JSON les KPIs accessibles selon le rôle/direction de l'utilisateur.
    - Admin / Validateur → tous les KPIs ayant une URL
    - Collecteur → uniquement les KPIs de sa/ses direction(s) (User.structure)
    Utilisé par le menu déroulant 'Choix du KPI' dans header.html
    """
    GROUPE_LABELS = {
        'GT1': 'Gouvernance et Réformes',
        'GT2': 'Travail et Sécurité Sociale',
        'GT3': 'Fonction Publique',
    }

    # ── Récupérer la direction de l'utilisateur ──────────────────────────────
    user_directions = []
    if current_user.structure:
        user_directions = [d.strip().upper() for d in current_user.structure.split(',') if d.strip()]

    # ── Récupérer TOUS les KPIs en BDD ────────────────────────────────────────
    tous_kpis = KPI.query.order_by(KPI.code).all()

    groupes = {}
    for kpi in tous_kpis:
        code_upper = (kpi.code or '').upper()

        # Direction du KPI : d'abord dans KPI_META, sinon dans axe_strategique (BDD)
        meta           = KPI_META.get(code_upper, {})
        direction_kpi  = meta.get('direction', '') or (kpi.axe_strategique or '').strip().upper()

        # ── Filtre selon le rôle ───────────────────────────────────────────────
        if current_user.role == 'admin':
            # Admin → voit tout
            pass
        else:
            # Collecteur et Validateur → uniquement les KPIs de leur direction
            if not user_directions:
                continue   # aucune direction affectée → rien
            if direction_kpi and direction_kpi not in user_directions:
                continue   # KPI d'une autre direction → ignoré

        # ── Groupe thématique ──────────────────────────────────────────────────
        gt = meta.get('groupe', '')
        if not gt:
            # Déduire depuis le code (GT1PKI01 → GT1)
            if code_upper.startswith('GT1'):   gt = 'GT1'
            elif code_upper.startswith('GT2'): gt = 'GT2'
            elif code_upper.startswith('GT3'): gt = 'GT3'
            else:                              gt = 'GT1'

        url = f'/kpi/{(kpi.code or "").lower()}'
        groupes.setdefault(gt, {'label': GROUPE_LABELS.get(gt, gt), 'kpis': []})
        groupes[gt]['kpis'].append({
            'code':      kpi.code,
            'intitule':  kpi.indicateur,
            'url':       url,
            'direction': direction_kpi,
        })

    # Résultat trié GT1 → GT2 → GT3
    result = [groupes[g] for g in ['GT1', 'GT2', 'GT3'] if g in groupes]
    total  = sum(len(g['kpis']) for g in result)

    return jsonify({
        'role':      current_user.role,
        'structure': current_user.structure or '',
        'groupes':   result,
        'total':     total,
    })


# ── API JSON : liste des KPIs pour la page admin ──────────────────────────────
@bp.route('/admin/api/kpis/liste')
@login_required
@admin_required
def admin_liste_kpis():
    """
    Retourne la liste complète des KPIs en JSON pour la page admin/liste_kpis.html.
    Inclut le nombre d'années renseignées pour la barre de progression.
    """
    kpis    = KPI.query.order_by(KPI.code).all()
    groupes = {g.id: g.nom for g in GroupeThematique.query.all()}

    result = []
    for k in kpis:
        # Compter les années renseignées (valeur non nulle)
        nb_renseigne = db.session.execute(text(
            "SELECT COUNT(*) FROM kpis_annuels "
            "WHERE kpi_id = :id AND valeur_calculee IS NOT NULL AND valeur_calculee != 0"
        ), {'id': k.id}).scalar() or 0

        result.append({
            'id':          k.id,
            'code':        k.code or '',
            'indicateur':  k.indicateur or '',
            'groupe':      groupes.get(k.groupe_id, '—'),
            'unite':       k.unite or '—',
            'formule':     k.formule_type or '—',
            'url':         f'/kpi/{(k.code or "").lower()}',
            'nb_renseigne': nb_renseigne,
            'complet':     nb_renseigne >= 11,
        })

    return jsonify({'kpis': result, 'total': len(result)})


# ── API JSON : détail d'un KPI (pour le modal d'édition) ─────────────────────
@bp.route('/admin/api/kpis/<int:kpi_id>', methods=['GET'])
@login_required
@admin_required
def admin_kpi_detail_get(kpi_id):
    kpi = KPI.query.get_or_404(kpi_id)
    return jsonify({
        'id':                kpi.id,
        'code':              kpi.code,
        'groupe_id':         kpi.groupe_id,
        'indicateur':        kpi.indicateur,
        'definition':        kpi.definition,
        'objectif':          kpi.objectif,
        'formule_type':      kpi.formule_type,
        'unite':             kpi.unite,
        'numerateur_label':  kpi.numerateur_label,
        'denominateur_label':kpi.denominateur_label,
        'frequence_collecte':kpi.frequence_collecte,
        'source_verification':kpi.source_verification,
        'valeur_cible':      kpi.valeur_cible,
        'axe_strategique':   kpi.axe_strategique,
        'direction':         kpi.axe_strategique or '',
    })


# ── API JSON : modification d'un KPI ─────────────────────────────────────────
@bp.route('/admin/api/kpis/<int:kpi_id>', methods=['PUT'])
@login_required
@admin_required
@log_action('UPDATE_KPI_META')
def admin_kpi_detail_put(kpi_id):
    kpi  = KPI.query.get_or_404(kpi_id)
    data = request.get_json(silent=True) or {}
    try:
        kpi.indicateur         = data.get('indicateur',         kpi.indicateur)
        kpi.definition         = data.get('definition',         kpi.definition)
        kpi.objectif           = data.get('objectif',           kpi.objectif)
        kpi.formule_type       = data.get('formule_type',       kpi.formule_type)
        kpi.unite              = data.get('unite',              kpi.unite)
        kpi.numerateur_label   = data.get('numerateur_label',   kpi.numerateur_label)
        kpi.denominateur_label = data.get('denominateur_label', kpi.denominateur_label)
        kpi.frequence_collecte = data.get('frequence_collecte', kpi.frequence_collecte)
        kpi.source_verification= data.get('source_verification',kpi.source_verification)
        kpi.axe_strategique    = data.get('axe_strategique',    kpi.axe_strategique)
        if data.get('valeur_cible') is not None:
            kpi.valeur_cible = float(data['valeur_cible'])
        if data.get('groupe_id'):
            kpi.groupe_id = int(data['groupe_id'])
        db.session.commit()
        return jsonify({'status': 'success', 'code': kpi.code})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ── API JSON : création d'un KPI ──────────────────────────────────────────────
@bp.route('/admin/api/kpis/create', methods=['POST'])
@login_required
@admin_required
@log_action('CREATE_KPI')
def admin_create_kpi():
    data = request.get_json(silent=True) or {}
    try:
        code       = (data.get('code') or '').strip().upper()
        indicateur = (data.get('indicateur') or '').strip()

        if not code:
            return jsonify({'error': 'Le code est obligatoire.'}), 400
        if not indicateur:
            return jsonify({'error': "L'intitulé est obligatoire."}), 400

        # Résoudre groupe_id
        groupe_id_raw = data.get('groupe_id', '')
        groupe_id = None
        if str(groupe_id_raw).isdigit():
            groupe_id = int(groupe_id_raw)
        elif groupe_id_raw:
            g = GroupeThematique.query.filter(
                db.func.lower(GroupeThematique.nom).contains(str(groupe_id_raw).lower())
            ).first() or GroupeThematique.query.first()
            if g:
                groupe_id = g.id
        if not groupe_id:
            return jsonify({'error': 'Groupe thématique invalide.'}), 400

        # Vérifier unicité
        if KPI.query.filter(db.func.upper(KPI.code) == code).first():
            return jsonify({'error': f'Le code {code} existe déjà.'}), 409

        # La direction est stockée dans axe_strategique en BDD
        direction = (data.get('direction') or data.get('axe_strategique') or '').strip().upper()

        kpi = KPI(
            code               = code,
            indicateur         = indicateur,
            groupe_id          = groupe_id,
            definition         = data.get('definition', ''),
            objectif           = data.get('objectif', ''),
            formule_type       = data.get('formule_type', 'pourcentage'),
            unite              = data.get('unite', '%'),
            numerateur_label   = data.get('numerateur_label', 'Numérateur'),
            denominateur_label = data.get('denominateur_label', 'Dénominateur'),
            frequence_collecte = data.get('frequence_collecte', 'Annuelle'),
            source_verification= data.get('source_verification', ''),
            axe_strategique    = direction,   # direction stockée ici
            valeur_cible       = float(data['valeur_cible']) if data.get('valeur_cible') else None,
        )
        db.session.add(kpi)
        db.session.flush()

        # Valeurs annuelles pré-remplies
        for annee_str, valeur in (data.get('valeurs_annuelles') or {}).items():
            if valeur is None:
                continue
            db.session.execute(text("""
                INSERT INTO kpis_annuels
                    (kpi_id, annee, numerateur_valeur, denominateur_valeur,
                     valeur_calculee, commentaire, created_at, created_by)
                VALUES (:kpi_id, :annee, :val, 0, :val, 'Saisie initiale', :now, :uid)
                ON CONFLICT(kpi_id, annee) DO UPDATE SET
                    valeur_calculee   = excluded.valeur_calculee,
                    numerateur_valeur = excluded.numerateur_valeur
            """), {
                'kpi_id': kpi.id, 'annee': int(annee_str),
                'val': float(valeur), 'now': datetime.utcnow(), 'uid': current_user.id
            })

        db.session.commit()
        return jsonify({
            'status':  'success',
            'id':      kpi.id,
            'code':    code,
            'url':     f'/kpi/{code.lower()}',
            'message': f'KPI {code} créé avec succès.',
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ── API JSON : suppression d'un KPI ──────────────────────────────────────────
@bp.route('/admin/api/kpis/<int:kpi_id>', methods=['DELETE'])
@login_required
@admin_required
@log_action('DELETE_KPI')
def admin_delete_kpi(kpi_id):
    try:
        kpi = KPI.query.get_or_404(kpi_id)
        code = kpi.code
        db.session.execute(text("DELETE FROM kpis_annuels WHERE kpi_id = :id"), {'id': kpi_id})
        db.session.execute(text("DELETE FROM preuves_documentaires WHERE kpi_id = :id"), {'id': kpi_id})
        db.session.execute(text("DELETE FROM commentaires_annuels WHERE kpi_id = :id"), {'id': kpi_id})
        db.session.delete(kpi)
        db.session.commit()
        return jsonify({'status': 'success', 'message': f'KPI {code} supprimé.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/admin/export/base-donnees')
@login_required
@admin_required
@log_action('EXPORT_EXCEL')
def export_base_donnees():
    """
    Génère et télécharge le fichier Excel BASE_DONNÉES complet,
    au même format que la feuille BASE_DONNÉES du classeur original MTFP.
    Accessible depuis le dashboard admin via le bouton Export.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    import datetime
 
    # -----------------------------------------------------------------------
    # 1. REQUÊTE BASE DE DONNÉES
    #    On récupère tous les KPIs avec leurs valeurs annuelles (2016→2026)
    #    et le nom du groupe thématique — en une seule requête SQL.
    # -----------------------------------------------------------------------
    rows = db.session.execute(text("""
        SELECT
            k.id,
            k.code,
            g.nom          AS groupe,
            k.indicateur,
            k.unite,
            ka_2016.valeur_calculee AS v2016,
            ka_2017.valeur_calculee AS v2017,
            ka_2018.valeur_calculee AS v2018,
            ka_2019.valeur_calculee AS v2019,
            ka_2020.valeur_calculee AS v2020,
            ka_2021.valeur_calculee AS v2021,
            ka_2022.valeur_calculee AS v2022,
            ka_2023.valeur_calculee AS v2023,
            ka_2024.valeur_calculee AS v2024,
            ka_2025.valeur_calculee AS v2025,
            ka_2026.valeur_calculee AS v2026
        FROM kpis k
        LEFT JOIN groupes_thematiques g ON g.id = k.groupe_id
        LEFT JOIN kpis_annuels ka_2016 ON ka_2016.kpi_id = k.id AND ka_2016.annee = 2016
        LEFT JOIN kpis_annuels ka_2017 ON ka_2017.kpi_id = k.id AND ka_2017.annee = 2017
        LEFT JOIN kpis_annuels ka_2018 ON ka_2018.kpi_id = k.id AND ka_2018.annee = 2018
        LEFT JOIN kpis_annuels ka_2019 ON ka_2019.kpi_id = k.id AND ka_2019.annee = 2019
        LEFT JOIN kpis_annuels ka_2020 ON ka_2020.kpi_id = k.id AND ka_2020.annee = 2020
        LEFT JOIN kpis_annuels ka_2021 ON ka_2021.kpi_id = k.id AND ka_2021.annee = 2021
        LEFT JOIN kpis_annuels ka_2022 ON ka_2022.kpi_id = k.id AND ka_2022.annee = 2022
        LEFT JOIN kpis_annuels ka_2023 ON ka_2023.kpi_id = k.id AND ka_2023.annee = 2023
        LEFT JOIN kpis_annuels ka_2024 ON ka_2024.kpi_id = k.id AND ka_2024.annee = 2024
        LEFT JOIN kpis_annuels ka_2025 ON ka_2025.kpi_id = k.id AND ka_2025.annee = 2025
        LEFT JOIN kpis_annuels ka_2026 ON ka_2026.kpi_id = k.id AND ka_2026.annee = 2026
        ORDER BY g.nom, k.code
    """)).fetchall()
 
    # -----------------------------------------------------------------------
    # 2. COULEURS (identiques à l'original Excel MTFP)
    # -----------------------------------------------------------------------
    C_ORANGE      = "C06000"   # Titre principal + en-têtes colonnes
    C_BLUE_DARK   = "344D6E"   # Titre section synthèse
    C_BLUE_MED    = "2E6DA4"   # Sous-titre + headers synthèse
    C_BLUE_LIGHT  = "D6E4F0"   # Lignes impaires + GT1 synthèse
    C_YELLOW      = "FFF2CC"   # Colonne Tendance
    C_GREY_LIGHT  = "F2F2F2"   # Colonne Unité
    C_ORANGE_LIGHT = "F4D6C8"  # GT2 synthèse
    C_GREEN_LIGHT  = "E2EFDA"  # GT3 synthèse
    C_WHITE        = "FFFFFF"
 
    def _fill(h):
        return PatternFill("solid", fgColor=h)
 
    def _font(bold=False, size=10, color="000000", italic=False):
        return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
 
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
 
    def _border():
        s = Side(border_style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
 
    # -----------------------------------------------------------------------
    # 3. CONSTRUCTION DU CLASSEUR EXCEL
    # -----------------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "BASE_DONNÉES"
 
    YEARS = [2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025, 2026]
 
    # --- Largeurs des colonnes (identiques à l'original) ---
    for col, w in {"A": 4, "B": 12, "C": 22, "D": 40,
                   "E": 11, "F": 11, "G": 11, "H": 11, "I": 11,
                   "J": 11, "K": 11, "L": 11, "M": 11, "N": 11,
                   "O": 11, "P": 14, "Q": 12, "R": 9}.items():
        ws.column_dimensions[col].width = w
 
    # --- Ligne 1-2 : Titre principal orange ---
    ws.row_dimensions[1].height = 21.75
    ws.row_dimensions[2].height = 15.0
    ws.merge_cells("A1:R2")
    c = ws["A1"]
    c.value = "MINISTÈRE DU TRAVAIL ET DE LA FONCTION PUBLIQUE — BILAN DÉCENNAL 2016-2026"
    c.font = _font(bold=True, size=13, color="FFFFFF")
    c.fill = _fill(C_ORANGE)
    c.alignment = _align("center", "center")
 
    # --- Ligne 3-4 : Sous-titre bleu ---
    ws.row_dimensions[3].height = 21.75
    ws.row_dimensions[4].height = 15.0
    ws.merge_cells("A3:R4")
    c = ws["A3"]
    c.value = (
        "BASE DE DONNÉES CONSOLIDÉE — INDICATEURS DE PERFORMANCE "
        "(Données exportées automatiquement depuis la base de données)"
    )
    c.font = _font(bold=True, size=11, color="FFFFFF")
    c.fill = _fill(C_BLUE_MED)
    c.alignment = _align("center", "center", wrap=True)
 
    ws.row_dimensions[5].height = 7.5  # ligne vide séparatrice
 
    # --- Ligne 6 : En-têtes colonnes ---
    ws.row_dimensions[6].height = 34.5
    headers = (
        ["N°", "Code", "Groupe Thématique", "Intitulé Indicateur"]
        + [str(y) for y in YEARS]
        + ["Moyenne\n2016-2026", "Tendance", "Unité"]
    )
    for ci, h in enumerate(headers, 1):
        c = ws.cell(6, ci, h)
        c.font  = _font(bold=True, size=9, color="FFFFFF")
        c.fill  = _fill(C_ORANGE)
        c.alignment = _align("center", "center", wrap=True)
        c.border = _border()
 
    # -----------------------------------------------------------------------
    # 4. LIGNES DE DONNÉES
    # -----------------------------------------------------------------------
    # Correspondance nom BDD → nom affiché + couleur synthèse
    GROUPES_META = {
        "Gouvernance et Reformes":    ("GT1 – Gouvernance et Réformes",    C_BLUE_LIGHT),
        "Travail et Securite Sociale": ("GT2 – Travail et Sécurité Sociale", C_ORANGE_LIGHT),
        "Fonction Publique":           ("GT3 – Fonction Publique",           C_GREEN_LIGHT),
    }
 
    # Calcul de la tendance à partir des valeurs disponibles
    def _tendance(valeurs):
        vals = [v for v in valeurs if v is not None]
        if len(vals) < 2:
            return "→ Stable"
        diff = vals[-1] - vals[0]
        if diff > 0:
            return "↑ Hausse"
        elif diff < 0:
            return "↓ Baisse"
        return "→ Stable"
 
    groupes_stats = {}   # pour la synthèse en bas
 
    for idx, row in enumerate(rows):
        kpi_id, code, groupe_nom, intitule, unite = row[0], row[1], row[2] or "", row[3], row[4] or ""
        vals = [row[5 + i] for i in range(11)]   # v2016 … v2026
 
        # Calcul moyenne (ignore les None)
        vals_ok = [v for v in vals if v is not None]
        moyenne  = round(sum(vals_ok) / len(vals_ok), 4) if vals_ok else None
        tendance = _tendance(vals)
        numero   = idx + 1
 
        # Stats pour la synthèse
        if groupe_nom not in groupes_stats:
            groupes_stats[groupe_nom] = {"total": 0, "hausse": 0, "baisse": 0, "stable": 0}
        groupes_stats[groupe_nom]["total"] += 1
        if "Hausse" in tendance:
            groupes_stats[groupe_nom]["hausse"] += 1
        elif "Baisse" in tendance:
            groupes_stats[groupe_nom]["baisse"] += 1
        else:
            groupes_stats[groupe_nom]["stable"] += 1
 
        data_row = 7 + idx
        ws.row_dimensions[data_row].height = 18.75
        row_fill = C_BLUE_LIGHT if idx % 2 == 0 else C_WHITE
 
        # Colonnes fixes
        for ci, val, bold, align_h, wrap in [
            (1, numero,    True,  "center", False),
            (2, code,      True,  "center", False),
            (3, groupe_nom, False, "left",  False),
            (4, intitule,  False, "left",   True),
        ]:
            c = ws.cell(data_row, ci, val)
            c.font  = _font(bold=bold, size=9)
            c.fill  = _fill(row_fill)
            c.alignment = _align(align_h, "center", wrap)
            c.border = _border()
 
        # Colonnes années (E=col5 → O=col15)
        for yi, val in enumerate(vals):
            c = ws.cell(data_row, 5 + yi, val)
            c.font  = _font(size=9)
            c.fill  = _fill(C_WHITE)
            c.alignment = _align("center", "center")
            c.border = _border()
            if val is not None:
                c.number_format = "#,##0.00"
 
        # Colonne Moyenne (P=col16)
        c = ws.cell(data_row, 16, moyenne)
        c.font  = _font(size=9)
        c.fill  = _fill(C_WHITE)
        c.alignment = _align("center", "center")
        c.border = _border()
        if moyenne is not None:
            c.number_format = "#,##0.00"
 
        # Colonne Tendance (Q=col17)
        c = ws.cell(data_row, 17, tendance)
        c.font  = _font(bold=True, size=9)
        c.fill  = _fill(C_YELLOW)
        c.alignment = _align("center", "center")
        c.border = _border()
 
        # Colonne Unité (R=col18)
        c = ws.cell(data_row, 18, unite)
        c.font  = _font(size=9)
        c.fill  = _fill(C_GREY_LIGHT)
        c.alignment = _align("center", "center")
        c.border = _border()
 
    # -----------------------------------------------------------------------
    # 5. TABLEAU DE SYNTHÈSE (en bas des données)
    # -----------------------------------------------------------------------
    last_data_row = 7 + len(rows)
    sep_row        = last_data_row + 1
    title_row      = sep_row + 1
    header_row     = sep_row + 2
    data_start_row = sep_row + 3
 
    ws.row_dimensions[sep_row].height = 7.5
 
    # Titre synthèse
    ws.row_dimensions[title_row].height = 21.75
    ws.merge_cells(f"A{title_row}:R{title_row}")
    c = ws[f"A{title_row}"]
    c.value = "TABLEAU DE SYNTHÈSE STATISTIQUE PAR GROUPE THÉMATIQUE"
    c.font  = _font(bold=True, size=11, color="FFFFFF")
    c.fill  = _fill(C_BLUE_DARK)
    c.alignment = _align("center", "center")
 
    # En-têtes synthèse
    ws.row_dimensions[header_row].height = 30
    synth_headers = [
        "Groupe Thématique", "Nb Indicateurs",
        "Indicateurs en Hausse ↑", "Indicateurs en Baisse ↓",
        "Indicateurs Stables →", "% Amélioration"
    ]
    for ci, h in enumerate(synth_headers, 1):
        c = ws.cell(header_row, ci, h)
        c.font  = _font(bold=True, size=9, color="FFFFFF")
        c.fill  = _fill(C_BLUE_MED)
        c.alignment = _align("center", "center", wrap=True)
        c.border = _border()
 
    # Données synthèse — dans l'ordre GT1, GT2, GT3
    ordre_groupes = [
        "Gouvernance et Reformes",
        "Travail et Securite Sociale",
        "Fonction Publique",
    ]
    for g_idx, g_key in enumerate(ordre_groupes):
        r = data_start_row + g_idx
        ws.row_dimensions[r].height = 18.75
        g_data   = groupes_stats.get(g_key, {"total": 0, "hausse": 0, "baisse": 0, "stable": 0})
        label_gt, g_fill = GROUPES_META.get(g_key, (g_key, C_WHITE))
        total    = g_data["total"]
        pct      = round(g_data["hausse"] / total * 100, 2) if total > 0 else 0
 
        for ci, val, bold in [
            (1, label_gt,          True),
            (2, total,             False),
            (3, g_data["hausse"],  False),
            (4, g_data["baisse"],  False),
            (5, g_data["stable"],  False),
            (6, pct,               False),
        ]:
            c = ws.cell(r, ci, val)
            c.font  = _font(bold=bold, size=9)
            c.fill  = _fill(g_fill)
            c.alignment = _align("left" if ci == 1 else "center", "center")
            c.border = _border()
            if ci == 6:
                c.number_format = '0.00"%"'
 
    # Pied de page horodaté
    footer_row = data_start_row + len(ordre_groupes) + 1
    ws.row_dimensions[footer_row].height = 15
    ws.merge_cells(f"A{footer_row}:R{footer_row}")
    now = datetime.datetime.now().strftime("%d/%m/%Y à %H:%M")
    c = ws[f"A{footer_row}"]
    c.value = (
        f"Exporté le {now} par {current_user.username} "
        f"— Bilan Décennal MTFP 2016-2026 | Plateforme de suivi des indicateurs"
    )
    c.font      = _font(size=8, italic=True, color="888888")
    c.alignment = _align("right", "center")
 
    # Volets figés : on fige les 4 premières colonnes + les 6 premières lignes
    ws.freeze_panes = "E7"
 
    # -----------------------------------------------------------------------
    # 6. ENVOI DU FICHIER
    # -----------------------------------------------------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
 
    filename = datetime.datetime.now().strftime("MTFP_Base_Donnees_%Y%m%d_%H%M.xlsx")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@bp.route('/admin/preview/base-donnees')
@login_required
@admin_required
def preview_base_donnees():
    """
    Retourne un JSON avec les données de la BASE_DONNÉES
    pour alimenter le panneau de prévisualisation.
    """
    rows = db.session.execute(text("""
        SELECT
            k.id, k.code, g.nom AS groupe, k.indicateur, k.unite,
            ka_2016.valeur_calculee, ka_2017.valeur_calculee, ka_2018.valeur_calculee,
            ka_2019.valeur_calculee, ka_2020.valeur_calculee, ka_2021.valeur_calculee,
            ka_2022.valeur_calculee, ka_2023.valeur_calculee, ka_2024.valeur_calculee,
            ka_2025.valeur_calculee, ka_2026.valeur_calculee
        FROM kpis k
        LEFT JOIN groupes_thematiques g ON g.id = k.groupe_id
        LEFT JOIN kpis_annuels ka_2016 ON ka_2016.kpi_id = k.id AND ka_2016.annee = 2016
        LEFT JOIN kpis_annuels ka_2017 ON ka_2017.kpi_id = k.id AND ka_2017.annee = 2017
        LEFT JOIN kpis_annuels ka_2018 ON ka_2018.kpi_id = k.id AND ka_2018.annee = 2018
        LEFT JOIN kpis_annuels ka_2019 ON ka_2019.kpi_id = k.id AND ka_2019.annee = 2019
        LEFT JOIN kpis_annuels ka_2020 ON ka_2020.kpi_id = k.id AND ka_2020.annee = 2020
        LEFT JOIN kpis_annuels ka_2021 ON ka_2021.kpi_id = k.id AND ka_2021.annee = 2021
        LEFT JOIN kpis_annuels ka_2022 ON ka_2022.kpi_id = k.id AND ka_2022.annee = 2022
        LEFT JOIN kpis_annuels ka_2023 ON ka_2023.kpi_id = k.id AND ka_2023.annee = 2023
        LEFT JOIN kpis_annuels ka_2024 ON ka_2024.kpi_id = k.id AND ka_2024.annee = 2024
        LEFT JOIN kpis_annuels ka_2025 ON ka_2025.kpi_id = k.id AND ka_2025.annee = 2025
        LEFT JOIN kpis_annuels ka_2026 ON ka_2026.kpi_id = k.id AND ka_2026.annee = 2026
        ORDER BY g.nom, k.code
    """)).fetchall()
 
    GROUPES_LABELS = {
        "Gouvernance et Reformes":     "GT1 – Gouvernance et Réformes",
        "Travail et Securite Sociale": "GT2 – Travail et Sécurité Sociale",
        "Fonction Publique":           "GT3 – Fonction Publique",
    }
 
    indicateurs = []
    groupes_stats = {}
    total_renseignes = 0
 
    for row in rows:
        vals = [row[5 + i] for i in range(11)]
        vals_ok = [v for v in vals if v is not None]
        moyenne = round(sum(vals_ok) / len(vals_ok), 2) if vals_ok else None
        if vals_ok:
            total_renseignes += 1
 
        diff = (vals_ok[-1] - vals_ok[0]) if len(vals_ok) >= 2 else 0
        if diff > 0:   tendance = "↑ Hausse"
        elif diff < 0: tendance = "↓ Baisse"
        else:          tendance = "→ Stable"
 
        groupe_nom = row[2] or ""
        if groupe_nom not in groupes_stats:
            groupes_stats[groupe_nom] = {"total": 0, "hausse": 0, "baisse": 0, "stable": 0, "renseignes": 0}
        groupes_stats[groupe_nom]["total"] += 1
        if vals_ok: groupes_stats[groupe_nom]["renseignes"] += 1
        if "Hausse" in tendance: groupes_stats[groupe_nom]["hausse"] += 1
        elif "Baisse" in tendance: groupes_stats[groupe_nom]["baisse"] += 1
        else: groupes_stats[groupe_nom]["stable"] += 1
 
        indicateurs.append({
            "numero":   len(indicateurs) + 1,
            "code":     row[1],
            "groupe":   GROUPES_LABELS.get(groupe_nom, groupe_nom),
            "intitule": row[3],
            "unite":    row[4] or "",
            "vals":     [round(v, 2) if v is not None else None for v in vals],
            "moyenne":  moyenne,
            "tendance": tendance,
        })
 
    groupes_list = []
    for g_key, g_label in GROUPES_LABELS.items():
        s = groupes_stats.get(g_key, {"total": 0, "hausse": 0, "baisse": 0, "stable": 0, "renseignes": 0})
        groupes_list.append({
            "label":      g_label,
            "total":      s["total"],
            "renseignes": s["renseignes"],
            "hausse":     s["hausse"],
            "baisse":     s["baisse"],
            "stable":     s["stable"],
            "pct_amelio": round(s["hausse"] / s["total"] * 100, 1) if s["total"] > 0 else 0,
        })
 
    return jsonify({
        "total":           len(indicateurs),
        "total_renseignes": total_renseignes,
        "groupes":         groupes_list,
        "indicateurs":     indicateurs,
        "annees":          list(range(2016, 2027)),
        "export_url":      url_for('api.export_base_donnees'),
    })


# ============ ADMIN - GESTION DES UTILISATEURS ============
@bp.route('/admin/utilisateurs')
@login_required
@admin_required
@log_action('VIEW_USERS')
def admin_utilisateurs():
    return render_template('admin/gestion_utilisateurs.html', active_page='admin_utilisateurs')


@bp.route('/admin/api/utilisateurs')
@login_required
@admin_required
def admin_get_users():
    users = User.query.all()
    return jsonify([{
        'id': u.id,
        'username': u.username,
        'email': u.email,
        'role': u.role,
        'structure': u.structure,
        'is_active': getattr(u, 'is_active', True),
        'created_at': u.created_at.isoformat() if u.created_at else None
    } for u in users])


@bp.route('/admin/api/utilisateurs', methods=['POST'])
@login_required
@admin_required
@log_action('CREATE_USER')
def admin_create_user():
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    data = request.json
    
    if User.query.filter_by(username=data['username']).first():
        return jsonify({'error': 'Ce nom d\'utilisateur existe déjà'}), 400
    
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'error': 'Cet email existe déjà'}), 400
    
    user = User(
        username=data['username'],
        email=data['email'],
        role=data.get('role', 'collecteur'),
        structure=data.get('structure', ''),
        is_active=data.get('is_active', True)
    )
    
    if data.get('password'):
        user.password = generate_password_hash(data['password'])
    else:
        user.password = generate_password_hash('password123')
    
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'success': True, 'id': user.id})


@bp.route('/admin/api/utilisateurs/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
@log_action('UPDATE_USER')
def admin_update_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    user = User.query.get_or_404(user_id)
    data = request.json
    
    user.username = data.get('username', user.username)
    user.email = data.get('email', user.email)
    user.role = data.get('role', user.role)
    user.structure = data.get('structure', user.structure)
    
    if 'is_active' in data:
        user.is_active = data['is_active']
    
    if data.get('password'):
        user.password = generate_password_hash(data['password'])
    
    db.session.commit()
    
    return jsonify({'success': True})


@bp.route('/admin/api/utilisateurs/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
@log_action('DELETE_USER')
def admin_delete_user(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    user = User.query.get_or_404(user_id)
    
    if user.id == current_user.id:
        return jsonify({'error': 'Vous ne pouvez pas supprimer votre propre compte'}), 400
    
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({'success': True})


@bp.route('/admin/api/utilisateurs/<int:user_id>/reset-password', methods=['POST'])
@login_required
@admin_required
@log_action('RESET_PASSWORD')
def admin_reset_password(user_id):
    if current_user.role != 'admin':
        return jsonify({'error': 'Non autorisé'}), 403
    
    user = User.query.get_or_404(user_id)
    data = request.json
    
    if not data.get('password'):
        return jsonify({'error': 'Mot de passe requis'}), 400
    
    user.password = generate_password_hash(data['password'])
    db.session.commit()
    
    return jsonify({'success': True})


# ============ ADMIN - LOGS ============
@bp.route('/admin/logs')
@login_required
@admin_required
@log_action('VIEW_LOGS')
def admin_logs():
    return render_template('admin/logs_consultation.html', active_page='admin_logs')


@bp.route('/admin/api/logs')
@login_required
@admin_required
def admin_api_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    include_stats = request.args.get('include_stats', 'false') == 'true'

    # ⚠️ CORRECTION : Récupérer TOUS les logs, pas seulement ceux de l'admin
    query = Log.query

    # Appliquer les filtres seulement si spécifiés
    if request.args.get('user_id') and request.args.get('user_id') != '':
        query = query.filter(Log.user_id == request.args.get('user_id', type=int))
    if request.args.get('action') and request.args.get('action') != '':
        query = query.filter(Log.action == request.args.get('action'))
    if request.args.get('date_debut'):
        query = query.filter(Log.created_at >= request.args.get('date_debut'))
    if request.args.get('date_fin'):
        query = query.filter(Log.created_at <= request.args.get('date_fin') + ' 23:59:59')

    total = query.count()
    logs = query.order_by(Log.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    
    # Récupérer les noms ET les rôles des utilisateurs
    users = {u.id: {'username': u.username, 'role': u.role} for u in User.query.all()}

    stats = None
    if include_stats:
        tous_les_logs = query.with_entities(Log.action).all()
        modifications = sum(1 for l in tous_les_logs if l.action in ('UPDATE_KPI', 'UPDATE_USER'))
        exports = sum(1 for l in tous_les_logs if l.action in ('EXPORT_EXCEL', 'EXPORT_PDF'))
        uploads = sum(1 for l in tous_les_logs if l.action == 'UPLOAD_FILE')
        stats = {
            'total': total,
            'modifications': modifications,
            'exports': exports,
            'uploads': uploads
        }

    return jsonify({
        'logs': [{
            'id': l.id,
            'user_name': users.get(l.user_id, {}).get('username', 'Inconnu'),
            'user_role': users.get(l.user_id, {}).get('role', '?'),
            'action': l.action,
            'entity_type': l.entity_type,
            'entity_id': l.entity_id,
            'details': l.details,
            'ip_address': l.ip_address,
            'created_at': l.created_at.strftime('%Y-%m-%d %H:%M:%S') if l.created_at else None
        } for l in logs],
        'page': page,
        'per_page': per_page,
        'total': total,
        'total_pages': (total + per_page - 1) // per_page,
        'stats': stats
    })


@bp.route('/admin/api/logs/export')
@login_required
@admin_required
@log_action('EXPORT_LOGS')
def admin_api_export_logs():
    import csv
    from io import StringIO
    from flask import Response
    
    query = Log.query
    
    if request.args.get('user_id') and request.args.get('user_id') != '':
        query = query.filter(Log.user_id == request.args.get('user_id', type=int))
    if request.args.get('action') and request.args.get('action') != '':
        query = query.filter(Log.action == request.args.get('action'))
    if request.args.get('date_debut'):
        query = query.filter(Log.created_at >= request.args.get('date_debut'))
    if request.args.get('date_fin'):
        query = query.filter(Log.created_at <= request.args.get('date_fin') + ' 23:59:59')
    
    logs = query.order_by(Log.created_at.desc()).all()
    users = {u.id: u.username for u in User.query.all()}
    
    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(['ID', 'Date', 'Utilisateur', 'Action', 'Détails', 'IP'])
    
    for log in logs:
        writer.writerow([
            log.id,
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
            users.get(log.user_id, 'Inconnu'),
            log.action,
            log.details or '',
            log.ip_address or ''
        ])
    
    output.seek(0)
    return Response(output.getvalue(), mimetype='text/csv; charset=utf-8-sig', headers={'Content-Disposition': 'attachment; filename=logs_export.csv'})


@bp.route('/admin/validations')
@login_required
@admin_required
@log_action('VIEW_VALIDATIONS')
def admin_validations():
    return render_template('admin/validations.html', active_page='admin_validations')


@bp.route('/api/admin/validations/liste')
@login_required
@admin_required
def api_admin_validations_liste():
    statut_filtre = request.args.get('statut', 'en_attente').strip()
    groupe_filtre = request.args.get('groupe', '').strip()
    annee_filtre  = request.args.get('annee', '').strip()
    search        = request.args.get('q', '').strip().lower()

    rows = db.session.execute(text('''
        SELECT
            ka.id, k.code, k.indicateur, g.nom,
            ka.annee, ka.numerateur_valeur, ka.denominateur_valeur,
            ka.valeur_calculee, k.unite, k.valeur_cible, ka.commentaire,
            k.axe_strategique,
            u_col.username,
            COALESCE((SELECT v.statut FROM validations v
                WHERE v.kpi_code = k.code AND v.annee = ka.annee
                ORDER BY v.id DESC LIMIT 1), 'en_attente') AS statut,
            COALESCE((SELECT v.commentaire FROM validations v
                WHERE v.kpi_code = k.code AND v.annee = ka.annee
                ORDER BY v.id DESC LIMIT 1), '') AS note_val,
            COALESCE((SELECT u2.username FROM validations v2
                JOIN users u2 ON u2.id = v2.validateur_id
                WHERE v2.kpi_code = k.code AND v2.annee = ka.annee
                ORDER BY v2.id DESC LIMIT 1), '') AS validateur,
            COALESCE((SELECT v3.updated_at FROM validations v3
                WHERE v3.kpi_code = k.code AND v3.annee = ka.annee
                ORDER BY v3.id DESC LIMIT 1), ka.created_at) AS date_action,
            (SELECT COUNT(*) FROM preuves_documentaires p
                WHERE p.kpi_id = k.id AND p.annee = ka.annee) AS nb_preuves
        FROM kpis_annuels ka
        JOIN kpis k ON k.id = ka.kpi_id
        JOIN groupes_thematiques g ON g.id = k.groupe_id
        LEFT JOIN users u_col ON u_col.id = ka.created_by
        WHERE ka.valeur_calculee IS NOT NULL
        ORDER BY ka.annee DESC, k.code
    ''')).fetchall()

    result = []
    for r in rows:
        statut = (r[13] or 'en_attente').strip().lower()
        if statut not in ('en_attente','valide','relance','annule'):
            statut = 'en_attente'
        if statut_filtre and statut_filtre != 'tous' and statut != statut_filtre:
            continue
        if groupe_filtre and groupe_filtre.lower() not in (r[3] or '').lower():
            continue
        if annee_filtre and str(r[4]) != annee_filtre:
            continue
        if search and search not in (r[1] or '').lower() and search not in (r[2] or '').lower():
            continue
        valeur = r[7]; cible = r[9]
        taux = round(valeur / cible * 100, 1) if cible and cible > 0 and valeur else 0
        result.append({
            'ka_id': r[0], 'kpi_code': r[1], 'kpi_intitule': r[2],
            'groupe': r[3], 'annee': r[4],
            'numerateur': r[5], 'denominateur': r[6],
            'valeur': round(valeur, 2) if valeur is not None else None,
            'unite': r[8] or '', 'valeur_cible': cible,
            'taux_atteinte': taux, 'commentaire': r[10] or '',
            'direction': r[11] or '', 'collecteur': r[12] or '—',
            'statut': statut, 'note_validateur': r[14] or '',
            'validateur': r[15] or '',
            'date_action': str(r[16])[:16] if r[16] else '',
            'nb_preuves': r[17] or 0,
            'url_kpi': f'/kpi/{(r[1] or "").lower()}',
        })

    return jsonify({
        'items': result, 'total': len(result),
        'stats': {
            'en_attente': sum(1 for r in result if r['statut'] == 'en_attente'),
            'valide':     sum(1 for r in result if r['statut'] == 'valide'),
            'relance':    sum(1 for r in result if r['statut'] == 'relance'),
            'annule':     sum(1 for r in result if r['statut'] == 'annule'),
        }
    })


@bp.route('/api/admin/validations/action/<int:ka_id>', methods=['POST'])
@login_required
@admin_required
@log_action('VALIDATE_KPI')
def api_admin_validation_action(ka_id):
    data   = request.get_json(silent=True) or {}
    action = (data.get('action') or '').strip().lower()
    note   = (data.get('note') or '').strip()
    MAP    = {'valider':'valide','valide':'valide','relancer':'relance',
              'relance':'relance','annuler':'annule','annule':'annule'}
    statut = MAP.get(action)
    if not statut:
        return jsonify({'error': 'Action invalide'}), 400
    if not note:
        return jsonify({'error': 'Un commentaire est obligatoire'}), 400
    row = KPISAnnuel.query.get_or_404(ka_id)
    kpi = KPI.query.get_or_404(row.kpi_id)
    try:
        db.session.add(Validation(
            kpi_code=kpi.code, annee=row.annee,
            validateur_id=current_user.id, statut=statut,
            commentaire=note, ancienne_valeur=row.valeur_calculee,
            nouvelle_valeur=row.valeur_calculee,
        ))
        db.session.commit()
        return jsonify({'status': 'success', 'statut': statut})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/api/admin/validations/action-groupe', methods=['POST'])
@login_required
@admin_required
@log_action('VALIDATE_KPI_GROUPE')
def api_admin_validation_groupe():
    data   = request.get_json(silent=True) or {}
    ids    = data.get('ids', [])
    action = (data.get('action') or '').strip().lower()
    note   = (data.get('note') or 'Action groupée').strip()
    MAP    = {'valider':'valide','valide':'valide','relancer':'relance',
              'relance':'relance','annuler':'annule','annule':'annule'}
    statut = MAP.get(action)
    if not statut or not ids:
        return jsonify({'error': 'Action ou IDs invalides'}), 400
    ok = 0
    for ka_id in ids:
        try:
            row = KPISAnnuel.query.get(int(ka_id))
            if not row: continue
            kpi = KPI.query.get(row.kpi_id)
            if not kpi: continue
            db.session.add(Validation(
                kpi_code=kpi.code, annee=row.annee,
                validateur_id=current_user.id, statut=statut,
                commentaire=note, ancienne_valeur=row.valeur_calculee,
                nouvelle_valeur=row.valeur_calculee,
            ))
            ok += 1
        except Exception:
            continue
    db.session.commit()
    return jsonify({'status': 'success', 'traites': ok})


@bp.route('/api/admin/stats')
@login_required
@admin_required
def admin_stats():
    users_data = db.session.execute(text("""
        SELECT 
            role,
            COUNT(*) as count,
            COUNT(CASE WHEN created_at >= date('now', '-30 days') THEN 1 END) as nouveaux
        FROM users 
        GROUP BY role
    """))
    
    kpi_data = db.session.execute(text("""
        SELECT 
            k.code,
            k.indicateur,
            g.nom as groupe,
            COUNT(DISTINCT ka.annee) as annees_renseignees,
            MAX(ka.annee) as derniere_annee,
            MAX(ka.valeur_calculee) as derniere_valeur
        FROM kpis k
        LEFT JOIN kpis_annuels ka ON k.id = ka.kpi_id
        LEFT JOIN groupes_thematiques g ON k.groupe_id = g.id
        GROUP BY k.id
        ORDER BY g.nom, k.code
    """))
    
    activite_recente = db.session.execute(text("""
        SELECT 
            DATE(created_at) as jour,
            COUNT(*) as total
        FROM logs 
        WHERE created_at >= date('now', '-30 days')
        GROUP BY DATE(created_at)
        ORDER BY jour DESC
    """)).fetchall()
    
    taux_collecte = db.session.execute(text("""
        SELECT 
            g.nom as groupe,
            COUNT(DISTINCT k.id) as total_kpi,
            COUNT(DISTINCT CASE WHEN ka.id IS NOT NULL THEN k.id END) as kpi_renseignes
        FROM groupes_thematiques g
        JOIN kpis k ON k.groupe_id = g.id
        LEFT JOIN kpis_annuels ka ON ka.kpi_id = k.id AND ka.annee = 2026
        GROUP BY g.id
    """)).fetchall()
    
    return jsonify({
        'users': [{'role': r[0], 'count': r[1], 'nouveaux': r[2]} for r in users_data],
        'kpi': [{'code': r[0], 'indicateur': r[1], 'groupe': r[2], 'annees': r[3], 'derniere_annee': r[4], 'derniere_valeur': r[5]} for r in kpi_data],
        'activite_recente': [{'jour': r[0], 'total': r[1]} for r in activite_recente],
        'taux_collecte': [{'groupe': r[0], 'total': r[1], 'renseignes': r[2], 'taux': round(r[2]/r[1]*100, 1) if r[1] > 0 else 0} for r in taux_collecte]
    })


@bp.route('/admin/api/logs/delete', methods=['DELETE'])
@login_required
@admin_required
def admin_api_delete_logs():
    """Supprimer des logs par période et/ou utilisateur"""
    try:
        date_debut = request.args.get('date_debut')
        date_fin = request.args.get('date_fin')
        user_id = request.args.get('user_id', type=int)
        
        query = Log.query
        
        if date_debut:
            query = query.filter(Log.created_at >= date_debut)
        if date_fin:
            query = query.filter(Log.created_at <= date_fin + ' 23:59:59')
        if user_id:
            query = query.filter(Log.user_id == user_id)
        
        count = query.count()
        query.delete()
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'{count} log(s) supprimé(s)'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


#############################################
####Dashboard VALIDATEUR (vue par direction)#
#############################################
@bp.route('/validateur/dashboard')
@login_required
@log_action('VIEW_VALIDATOR')
def validateur_dashboard():
    if current_user.role not in ['validateur', 'admin']:
        flash('Accès réservé aux validateurs', 'danger')
        return redirect(url_for('api.dashboard'))
    return render_template('dashboard/validateur_dashboard.html', active_page='validateur_dashboard')


@bp.route('/api/validateur/kpis')
@login_required
def validateur_kpis():
    if current_user.role not in ['validateur', 'admin']:
        return jsonify({'error': 'Non autorisé'}), 403
    
    query = text("""
        SELECT 
            k.id, k.code, k.indicateur, g.nom as groupe,
            ka.annee, ka.valeur_calculee, ka.commentaire,
            k.valeur_cible, k.unite
        FROM kpis k
        JOIN groupes_thematiques g ON k.groupe_id = g.id
        LEFT JOIN kpis_annuels ka ON ka.kpi_id = k.id AND ka.annee = 2026
        WHERE g.nom = :direction OR :direction = 'admin'
        ORDER BY g.nom, k.code
    """)
    
    result = db.session.execute(query, {'direction': current_user.structure or ''})
    
    kpis = []
    for row in result:
        kpis.append({
            'id': row[0],
            'code': row[1],
            'indicateur': row[2],
            'groupe': row[3],
            'annee': row[4],
            'valeur': row[5],
            'commentaire': row[6],
            'cible': row[7],
            'unite': row[8],
            'taux_atteinte': round(row[5] / row[7] * 100, 1) if row[7] and row[7] > 0 else 0
        })
    
    return jsonify(kpis)


#############################################
####Dashboard COLLECTEUR#####################
#############################################
@bp.route('/collecteur/dashboard')
@login_required
@log_action('VIEW_COLLECTOR')
def collecteur_dashboard():
    if current_user.role != 'collecteur':
        flash('Accès réservé aux collecteurs', 'danger')
        return redirect(url_for('api.dashboard'))
    return render_template('dashboard/collecteur_dashboard.html', active_page='collecteur_dashboard')


@bp.route('/collecteur/recap-soumissions')
@login_required
@log_action('VIEW_COLLECTOR_RECAP')
def collecteur_recap_soumissions():
    if current_user.role != 'collecteur':
        flash('Accès réservé aux collecteurs', 'danger')
        return redirect(url_for('api.dashboard'))
    return render_template('collecteur/recap_soumission.html', active_page='collecteur_recap_soumissions')


@bp.route('/validateur/recap-soumissions')
@login_required
@log_action('VIEW_VALIDATOR_RECAP')
def validateur_recap_soumissions():
    if current_user.role != 'validateur':
        flash('Accès réservé aux validateur', 'danger')
        return redirect(url_for('api.dashboard'))
    return render_template('validateur/recap_soumission.html', active_page='validation_recap_soumissions')


@bp.route('/api/collecteur/recap-soumissions')
@login_required
def api_collecteur_recap_soumissions():
    if current_user.role != 'collecteur':
        return jsonify({'error': 'Non autorisé'}), 403

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    filters = {
        'code': request.args.get('code', '').strip(),
        'indicateur': request.args.get('indicateur', '').strip(),
        'annee': request.args.get('annee', '').strip(),
        'statut': request.args.get('statut', '').strip(),
    }

    try:
        # fonction _load_recap_rows() pour récup toutes les lignes d'indicateurs soumis par le collecteur
        # groupe_id=0 = tous les groupes
        groupe_id = request.args.get('groupe_id', 0, type=int)
        items = _load_recap_rows(groupe_id, filters)
        payload = _paginate_items(items, page, per_page)
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/validateur/recap-soumissions')
@login_required
def api_validateur_recap_soumissions():
    if current_user.role != 'validateur':
        return jsonify({'error': 'Non autorisé'}), 403

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    filters = {
        'code': request.args.get('code', '').strip(),
        'indicateur': request.args.get('indicateur', '').strip(),
        'annee': request.args.get('annee', '').strip(),
        'statut': request.args.get('statut', '').strip(),
    }

    try:
        groupe_id = request.args.get('groupe_id', 0, type=int)
        items = _load_recap_rows(groupe_id, filters)
        payload = _paginate_items(items, page, per_page)
        return jsonify(payload)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/api/validateur/recap-soumissions/action/<int:recap_id>', methods=['POST'])
@login_required
def api_validateur_recap_action(recap_id):
    if current_user.role != 'validateur':
        return jsonify({'error': 'Non autorisé'}), 403

    data = request.get_json(silent=True) or {}
    action = (data.get('action') or '').strip().lower()
    note = (data.get('note') or '').strip()

    action_map = {
        'valider': 'valide',
        'relancer': 'relance',
        'annuler': 'annule',
        'valide': 'valide',
        'relance': 'relance',
        'annule': 'annule',
    }
    action = action_map.get(action, action)

    if action not in {'valide', 'relance', 'annule'}:
        return jsonify({'error': 'Action non autorisée'}), 400

    if not note:
        return jsonify({'error': 'Une note est obligatoire'}), 400

    row = KPISAnnuel.query.get_or_404(recap_id)
    kpi = KPI.query.get_or_404(row.kpi_id)

    latest_validation = _load_latest_validation(kpi.code or '', row.annee)
    latest_status = (latest_validation[0] or '').lower() if latest_validation else ''
    if latest_status in {'valide', 'annule'}:
        return jsonify({'error': 'Cette soumission a déjà été validée ou annulée'}), 400

    try:
        validation = Validation(
            kpi_code=kpi.code,
            annee=row.annee,
            validateur_id=current_user.id,
            statut=action,
            commentaire=note,
            ancienne_valeur=row.valeur_calculee,
            nouvelle_valeur=row.valeur_calculee,
        )
        db.session.add(validation)

        commentaire = CommentaireAnnuel.query.filter_by(kpi_id=row.kpi_id, annee=row.annee).one_or_none()
        if commentaire:
            commentaire.commentaire = note
            commentaire.date_ajout = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
            commentaire.ajoute_par = current_user.username
        else:
            commentaire = CommentaireAnnuel(
                kpi_id=row.kpi_id,
                annee=row.annee,
                commentaire=note,
                date_ajout=datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                ajoute_par=current_user.username,
            )
            db.session.add(commentaire)

        row.commentaire = note
        db.session.commit()

        return jsonify({
            'status': 'success',
            'action': action,
            'recap_id': recap_id,
            'message': 'Action enregistrée',
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500