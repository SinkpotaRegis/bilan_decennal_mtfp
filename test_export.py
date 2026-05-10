from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

print("Test d'export Excel...")

try:
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "TEST"
    ws['A1'].font = Font(bold=True)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    print("✅ Export Excel réussi !")
    
except Exception as e:
    print(f"❌ Erreur: {e}")